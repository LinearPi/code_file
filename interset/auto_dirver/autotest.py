from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import time
from openpyxl import load_workbook

def login_shence():
    i = 0
    # 打开网页
    driver = webdriver.Chrome('/Users/gouweiqi/Documents/code_py/code_file/interset/auto_dirver/chromedriver') 
    driver.get('http://sensor.xiguacity.cn:8107/login/index.html?project=production')
    driver.set_window_size(1400,900)
    # 登陆网页
    time.sleep(2)
    name = driver.find_element_by_xpath('//*[@id="userName"]')
    name.send_keys('lijie@xigua.city')
    passwd = driver.find_element_by_xpath('//*[@id="password"]')
    passwd.send_keys("DPt5Az")
    time.sleep(1)
    loggin = driver.find_element_by_xpath('//*[@id="submit"]')
    loggin.click()
    time.sleep(3)
    # 登陆成功
    
    
    wb = load_workbook(filename=file, read_only=True)
    ws = wb['Sheet2']
    for row in ws.rows:
        # if row[6]:
        # 点击用户
        search_botton = driver.find_element_by_xpath('//*[@id="sa_head_nav"]/ul/li[1]/a/span[2]').click()
        time.sleep(1)
        # 清除数据
        driver.find_element_by_xpath('//*[@id="test"]').clear()
        time.sleep(0.5)
        # 输入id
        print(row[6].value)
        imputt_data = driver.find_element_by_xpath('//*[@id="test"]').send_keys(row[6].value)
        time.sleep(1)
        # 确认用户
        click_botton = driver.find_element_by_xpath('//*[@id="searchTool"]/div[2]/div[1]/span').click()
        time.sleep(1)
    
        driver.find_element_by_xpath('//*[@id="searchTool"]/div[2]/div[2]/div[3]/div/div/span[2]').click()
        time.sleep(1)
        # 修改日期
        change_date1 = driver.find_element_by_xpath('//*[@id="inputDate"]').click()
        
        time.sleep(1)
        change_date2 = driver.find_element_by_xpath('//*[@id="sa-body"]/div[13]/div[3]/div[3]/ul/li[6]').click()
        time.sleep(30)
            
        # # 截图
        # driver.get_screenshot_as_file("/Users/gouweiqi/Desktop/1314img/{}_{}.png".format(i, row[1].value))
        # i += 1

    driver.close()
    driver.quit()

# def take_img():
#     driver.get_screenshot_as_file("/Users/gouweiqi/Desktop/1314img/{}.png".format(12345))

#     print('1')


def read_data(file):
    wb = load_workbook(filename=file, read_only=True)
    ws = wb['Sheet3']
    for row in ws.rows:

        print(row[6].value)


if __name__=='__main__':
    file = r'/Users/gouweiqi/Desktop/数据/12-14及23-24未完课分析名单20181113.xlsx'
    login_shence()
    # take_img()
    # read_data(file)
  